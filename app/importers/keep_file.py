"""Keep 官方导出文件导入器（M4b，设计文档 §3.7 通道 3）。

CLI：python -m app.importers.keep_file <path> [--dry-run] [--password PWD]
Web 上传（settings 模块）BackgroundTasks 调 import_keep(path, job_id=..., password=...)。

支持三种输入：
- .7z（py7zr）：Keep 客服导出重压缩包。实测解包为 <uid>/ 目录：
  <uid>.xlsx（sheet「运动记录」）+ user<uid>.xlsx（用户信息 key-value）
  + bundle-<uid>.zip（.fit 全是 ~180B 占位 stub，跳过并在报告注明）
- .zip（pyzipper，兼容 AES 加密 + password，密码错误给明确报错）
- 裸 .xlsx（直接按运动记录解析）

归一化：
- 运动记录逐行 → import_raw(source='keep_file', record_type='exercise'，
  「心率记录/运动轨迹」两个大字段替换为长度摘要) → workout_logs(source='keep')。
  external_id = sha1(开始时间|运动类型|运动时长)；开始时间为北京时间（+08:00）。
- user*.xlsx 不入业务表，仅留档 import_raw record_type='profile'。
- 跨源去重：Keep 和三星手表会记录同一次运动（07-03 跑步实测两边都有，相差 31s），
  已存在 source IN ('samsung_zip','health_connect') 且 started_at 相差 ≤3 分钟
  → 跳过入库，import_raw 置 parse_status='skipped'/parse_error='dup:samsung'，
  报告单列 dup_samsung 计数。
"""
from __future__ import annotations

import argparse
import hashlib
import json
import lzma
import shutil
import tempfile
import zipfile
from bisect import bisect_left
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import py7zr
import pyzipper
from openpyxl import load_workbook
from sqlalchemy import literal_column, select
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.orm import Session
from sqlalchemy.sql import text as sql_text

from app.db import SessionLocal
from app.models import ImportJob, ImportRaw, SyncState, WorkoutLog
from app.timeutil import now_local

PARSER_VERSION = 1
SOURCE = "keep_file"        # import_raw / sync_state 的来源标识
WORKOUT_SOURCE = "keep"     # workout_logs.source（CHECK 枚举里已有）
DUP_SOURCES = ("samsung_zip", "health_connect")
DUP_TOLERANCE_S = 180       # 跨源去重窗口：started_at 相差 ≤3 分钟
BJT = timezone(timedelta(hours=8))  # Keep 导出时间为北京时间，无 DST
BATCH_SIZE = 800

# 「运动记录」表头 → 内部键（前缀匹配，容忍「(秒)/(米)」等单位后缀变体）
_HEADER_KEYS = {
    "运动类型": "type",
    "运动时长": "dur_s",
    "开始时间": "start",
    "结束时间": "end",
    "卡路里": "cal",
    "运动距离": "dist_m",
    "平均心率": "avg_hr",
    "最大心率": "max_hr",
    "心率记录": "hr_json",
    "运动轨迹": "track",
}


class KeepImportError(ValueError):
    """输入文件问题（密码错误/格式不支持/缺运动记录表），报错文案直接面向用户。"""


# ---------- 基础工具 ----------

def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    return str(v).strip()


def _fnum(v: Any) -> float | None:
    s = _cell_str(v)
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _pos(v: float | None) -> float | None:
    """0/负值视为无数据。"""
    return v if v is not None and v > 0 else None


def _parse_start(v: Any) -> datetime:
    """'2026-07-03 08:17:10'（北京时间）→ aware UTC datetime。"""
    dt = v if isinstance(v, datetime) else datetime.fromisoformat(_cell_str(v))
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=BJT)
    return dt.astimezone(timezone.utc)


def _dup_near(sorted_ts: list[datetime], ts: datetime) -> bool:
    """ts 与已有三星/HC 记录 started_at 相差 ≤3 分钟？（sorted_ts 已排序，二分）"""
    i = bisect_left(sorted_ts, ts)
    for j in (i - 1, i):
        if 0 <= j < len(sorted_ts):
            if abs((sorted_ts[j] - ts).total_seconds()) <= DUP_TOLERANCE_S:
                return True
    return False


def _job_update(job_id: int | None, **fields: Any) -> None:
    """import_jobs 进度用独立短会话立即提交，供 Web 轮询可见。"""
    if not job_id:
        return
    try:
        with SessionLocal() as s:
            job = s.get(ImportJob, job_id)
            if job is not None:
                for k, v in fields.items():
                    setattr(job, k, v)
                s.commit()
    except Exception:
        pass  # 进度写失败不影响导入本身


# ---------- 解包：.7z / .zip / 裸 .xlsx → workdir 内的 xlsx 清单 ----------

def _extract_archive(src: Path, workdir: Path, password: str | None, report: dict) -> None:
    suffix = src.suffix.lower()
    if suffix == ".7z":
        try:
            with py7zr.SevenZipFile(src, password=password or None) as z:
                z.extractall(path=workdir)
        except py7zr.exceptions.PasswordRequired as e:
            raise KeepImportError("7z 为加密包，需要解压密码（上传时填写）") from e
        except py7zr.Bad7zFile as e:
            raise KeepImportError(f"不是有效的 7z 文件：{e}") from e
        except lzma.LZMAError as e:
            raise KeepImportError("7z 解压失败：密码错误或文件损坏") from e
    elif suffix == ".zip":
        try:
            with pyzipper.AESZipFile(src) as zf:
                if password:
                    zf.setpassword(password.encode("utf-8"))
                zf.extractall(path=workdir)
        except RuntimeError as e:
            msg = str(e)
            if "password" in msg.lower():
                raise KeepImportError(
                    "zip 解压密码错误或缺失（Keep 客服导出的 zip 为 AES 加密，需在上传时填写密码）"
                ) from e
            raise
        except zipfile.BadZipFile as e:
            raise KeepImportError(f"不是有效的 zip 文件：{e}") from e
    else:
        raise KeepImportError(f"不支持的文件类型 {src.suffix}（仅支持 .7z / .zip / .xlsx）")

    # 嵌套 zip（bundle-*.zip）：.fit 占位 stub 只清点不导入；若内含 xlsx 则一并解出
    for zp in sorted(workdir.rglob("*.zip")):
        info: dict[str, Any] = {"name": zp.name}
        try:
            with zipfile.ZipFile(zp) as zf:
                members = zf.infolist()
                info["members"] = len(members)
                fits = [m for m in members if m.filename.lower().endswith(".fit")]
                if fits:
                    sizes = [m.file_size for m in fits]
                    info["fit_stubs_skipped"] = len(fits)
                    info["fit_size_bytes"] = [min(sizes), max(sizes)]
                    info["note"] = ".fit 全是占位 stub（Keep 导出不含真实轨迹二进制），跳过"
                for m in members:
                    if m.filename.lower().endswith(".xlsx"):
                        zf.extract(m, workdir / f"_nested_{zp.stem}")
        except Exception as e:  # noqa: BLE001 嵌套包问题不阻断主导入
            info["error"] = str(e)[:200]
        report["files"]["nested_zips"].append(info)


def _collect_xlsx(src: Path, workdir: Path, password: str | None, report: dict) -> list[Path]:
    if src.suffix.lower() == ".xlsx":
        return [src]
    _extract_archive(src, workdir, password, report)
    return sorted(workdir.rglob("*.xlsx"))


# ---------- import_raw / workout_logs 批量写入 ----------

def _raw_item(record_type: str, raw: dict, external_id: str,
              status: str = "parsed", error: str | None = None) -> dict:
    return {
        "source": SOURCE,
        "record_type": record_type,
        "external_id": external_id,
        "raw": raw,
        "time_offset": "UTC+0800",
        "parse_status": status,
        "parse_error": error,
        "parse_version": PARSER_VERSION,
        "last_seen_at": now_local(),
    }


def _flush_raw(db: Session | None, dry: bool, batch: list[dict], stats: dict) -> None:
    """ON CONFLICT (source,record_type,external_id) 只刷新 last_seen_at（≈DO NOTHING）。"""
    if not batch or dry:
        batch.clear()
        return
    seen: set[str] = set()
    uniq: list[dict] = []
    for item in batch:
        if item["external_id"] in seen:
            stats["raw_skipped"] = stats.get("raw_skipped", 0) + 1
            continue
        seen.add(item["external_id"])
        uniq.append(item)
    batch.clear()
    if not uniq:
        return
    now = now_local()
    ins = pg_insert(ImportRaw).values(uniq)
    stmt = ins.on_conflict_do_update(
        index_elements=["source", "record_type", "external_id"],
        set_={"last_seen_at": now},
    ).returning(literal_column("(xmax = 0)"))
    flags = [r[0] for r in db.execute(stmt)]
    ins_n = sum(1 for f in flags if f)
    stats["raw_inserted"] = stats.get("raw_inserted", 0) + ins_n
    stats["raw_skipped"] = stats.get("raw_skipped", 0) + (len(uniq) - ins_n)


def _flush_norm(db: Session | None, dry: bool, batch: list[dict], stats: dict) -> None:
    if not batch or dry:
        batch.clear()
        return
    ins = pg_insert(WorkoutLog).values(list(batch))
    batch.clear()
    stmt = ins.on_conflict_do_update(
        index_elements=["source", "external_id"],
        index_where=sql_text("external_id IS NOT NULL"),
        set_={
            "log_date": ins.excluded.log_date,
            "started_at": ins.excluded.started_at,
            "session_type": ins.excluded.session_type,
            "duration_min": ins.excluded.duration_min,
            "distance_km": ins.excluded.distance_km,
            "calories": ins.excluded.calories,
            "avg_hr": ins.excluded.avg_hr,
            "max_hr": ins.excluded.max_hr,
            "detail": ins.excluded.detail,
            "updated_at": sql_text("now()"),
        },
    ).returning(literal_column("(xmax = 0)"))
    flags = [r[0] for r in db.execute(stmt)]
    n = sum(1 for f in flags if f)
    stats["inserted"] += n
    stats["skipped"] += len(flags) - n


# ---------- 工作表定位与解析 ----------

def _find_exercise_sheet(wb) -> tuple[Any, dict[int, str]] | None:
    """返回 (worksheet, 列号→内部键)；按表头前缀匹配，找不到返回 None。"""
    for ws in wb.worksheets:
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            continue
        colmap: dict[int, str] = {}
        for i, cell in enumerate(header):
            name = _cell_str(cell)
            for prefix, key in _HEADER_KEYS.items():
                if name.startswith(prefix):
                    colmap[i] = key
                    break
        keys = set(colmap.values())
        if "type" in keys and "start" in keys:
            return ws, colmap
    return None


def _import_exercise(db: Session | None, dry: bool, wb_path: Path, dup_ts: list[datetime],
                     stats: dict, ranges: dict, job_id: int | None) -> bool:
    """单个业务 xlsx 的运动记录导入；找到运动记录表返回 True。"""
    wb = load_workbook(wb_path, read_only=True, data_only=True)
    try:
        found = _find_exercise_sheet(wb)
        if found is None:
            return False
        ws, colmap = found
        if ws.max_row and ws.max_row > 1:
            _job_update(job_id, total=ws.max_row - 1)
        raw_batch: list[dict] = []
        norm_batch: list[dict] = []
        seen_ids: set[str] = set()
        type_counts: dict[str, int] = {}

        for parts in ws.iter_rows(min_row=2, values_only=True):
            if parts is None or all(v is None or _cell_str(v) == "" for v in parts):
                continue
            stats["rows"] += 1
            row = {key: parts[i] for i, key in colmap.items() if i < len(parts)}
            type_s = _cell_str(row.get("type"))
            start_s = _cell_str(row.get("start"))
            dur_s = _cell_str(row.get("dur_s"))
            ext_id = hashlib.sha1(f"{start_s}|{type_s}|{dur_s}".encode("utf-8")).hexdigest()

            # 留档 raw：两个大字段换成长度摘要
            hr_s = _cell_str(row.get("hr_json"))
            n_samples = 0
            if hr_s:
                try:
                    arr = json.loads(hr_s)
                    n_samples = len(arr) if isinstance(arr, list) else 0
                except ValueError:
                    n_samples = 0
            track_s = _cell_str(row.get("track"))
            raw = {
                "运动类型": type_s,
                "运动时长(秒)": dur_s,
                "开始时间": start_s,
                "结束时间": _cell_str(row.get("end")),
                "卡路里": _cell_str(row.get("cal")),
                "运动距离(米)": _cell_str(row.get("dist_m")),
                "平均心率": _cell_str(row.get("avg_hr")),
                "最大心率": _cell_str(row.get("max_hr")),
                "心率记录": f"<{n_samples} samples, len {len(hr_s)}>",
                "运动轨迹": f"<len {len(track_s)}>",
            }

            try:
                started = _parse_start(row.get("start"))
                log_date = started.astimezone(BJT).date()
                if "min" not in ranges or log_date < ranges["min"]:
                    ranges["min"] = log_date
                if "max" not in ranges or log_date > ranges["max"]:
                    ranges["max"] = log_date
                type_counts[type_s or "?"] = type_counts.get(type_s or "?", 0) + 1

                if _dup_near(dup_ts, started):
                    # 跨源去重：三星手表已记录同一次运动
                    stats["dup_samsung"] += 1
                    raw_batch.append(_raw_item("exercise", raw, ext_id, "skipped", "dup:samsung"))
                elif ext_id in seen_ids:
                    stats["skipped"] += 1  # 文件内重复行
                    raw_batch.append(_raw_item("exercise", raw, ext_id))
                else:
                    seen_ids.add(ext_id)
                    dur = _fnum(row.get("dur_s"))
                    dist = _fnum(row.get("dist_m"))
                    cal = _fnum(row.get("cal"))
                    norm_batch.append({
                        "log_date": log_date,
                        "started_at": started,
                        "session_type": type_s or None,
                        "duration_min": round(dur / 60) if dur is not None else None,
                        "distance_km": round(dist / 1000, 2) if _pos(dist) else None,
                        "calories": round(cal) if _pos(cal) else None,
                        "avg_hr": (round(v) if (v := _pos(_fnum(row.get("avg_hr")))) else None),
                        "max_hr": (round(v) if (v := _pos(_fnum(row.get("max_hr")))) else None),
                        "detail": {"hr_samples": n_samples},
                        "source": WORKOUT_SOURCE,
                        "external_id": ext_id,
                        "notes": None,
                    })
                    raw_batch.append(_raw_item("exercise", raw, ext_id))
            except Exception as e:  # noqa: BLE001 单行失败不阻断
                stats["failed"] += 1
                raw_batch.append(_raw_item("exercise", raw, ext_id, "failed", str(e)[:500]))

            if len(raw_batch) >= BATCH_SIZE:
                _flush_raw(db, dry, raw_batch, stats)
            if len(norm_batch) >= BATCH_SIZE:
                _flush_norm(db, dry, norm_batch, stats)
        _flush_raw(db, dry, raw_batch, stats)
        _flush_norm(db, dry, norm_batch, stats)
        for k, v in type_counts.items():
            stats["session_type_counts"][k] = stats["session_type_counts"].get(k, 0) + v
        return True
    finally:
        wb.close()  # read_only 模式必须显式关闭，否则句柄占用导致临时目录清理失败


def _import_profile(db: Session | None, dry: bool, wb_path: Path, stats: dict) -> None:
    """user*.xlsx：key-value 用户信息，仅留档 import_raw(record_type='profile')。"""
    wb = load_workbook(wb_path, read_only=True, data_only=True)
    try:
        raw_batch: list[dict] = []
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            next(rows, None)  # 表头（名称/值）
            for parts in rows:
                if parts is None or all(v is None or _cell_str(v) == "" for v in parts):
                    continue
                stats["rows"] += 1
                key = _cell_str(parts[0]) if len(parts) > 0 else ""
                val = _cell_str(parts[1]) if len(parts) > 1 else ""
                ext_id = hashlib.sha1(f"{key}|{val}".encode("utf-8")).hexdigest()
                raw_batch.append(_raw_item(
                    "profile", {"名称": key, "值": val[:1000]}, ext_id, "skipped"))
                stats["skipped"] += 1
        _flush_raw(db, dry, raw_batch, stats)
    finally:
        wb.close()


# ---------- 公共入口（settings 模块 Web 上传调用同一函数） ----------

def import_keep(path: str | Path, db: Session | None = None, dry_run: bool = False,
                job_id: int | None = None, password: str | None = None) -> dict:
    """导入 Keep 导出文件（.7z / .zip / 裸 .xlsx）。

    返回报告 dict：{"tables": {"exercise": {"rows","inserted","skipped","dup_samsung",
    "failed",...}, "profile": {...}}, "date_ranges": ..., "files": ...}。
    dry_run 只解析出报告不写库（跨源去重仍查库比对）；job_id 非空时把进度/结果写
    health.import_jobs（独立会话即时提交，供 Web 轮询）。
    """
    path = Path(path)
    report: dict[str, Any] = {
        "dry_run": dry_run,
        "source_file": path.name,
        "files": {"workbooks": [], "profile_workbooks": [], "nested_zips": []},
        "tables": {},
        "date_ranges": {},
    }
    ex_stats = {"rows": 0, "inserted": 0, "skipped": 0, "dup_samsung": 0, "failed": 0,
                "session_type_counts": {}}
    pf_stats = {"rows": 0, "inserted": 0, "skipped": 0, "failed": 0,
                "note": "仅留档 import_raw(record_type='profile')，不入业务表"}
    own_session = db is None
    if own_session:
        db = SessionLocal()
    _job_update(job_id, status="running", started_at=now_local())
    workdir = Path(tempfile.mkdtemp(prefix="keep_import_"))
    try:
        xlsx_paths = _collect_xlsx(path, workdir, password, report)
        if not xlsx_paths:
            raise KeepImportError("压缩包内未找到任何 .xlsx 文件")
        profile_books = [p for p in xlsx_paths if p.name.lower().startswith("user")]
        biz_books = [p for p in xlsx_paths if p not in profile_books]
        report["files"]["workbooks"] = [p.name for p in biz_books]
        report["files"]["profile_workbooks"] = [p.name for p in profile_books]

        # 跨源去重时间线：一次取全三星/HC 的 started_at，逐行二分比对
        dup_ts = sorted(
            db.execute(
                select(WorkoutLog.started_at).where(
                    WorkoutLog.source.in_(DUP_SOURCES),
                    WorkoutLog.started_at.is_not(None),
                )
            ).scalars()
        )

        for p in profile_books:
            _import_profile(db, dry_run, p, pf_stats)
        report["tables"]["profile"] = pf_stats

        ranges: dict = {}
        found_any = False
        for p in biz_books:
            found_any = _import_exercise(db, dry_run, p, dup_ts, ex_stats, ranges, job_id) or found_any
        if not found_any:
            raise KeepImportError("未在导出文件中找到「运动记录」工作表")
        report["tables"]["exercise"] = ex_stats
        if "min" in ranges:
            report["date_ranges"]["exercise"] = [ranges["min"].isoformat(),
                                                 ranges["max"].isoformat()]

        if not dry_run:
            ins = pg_insert(SyncState).values(
                source=SOURCE, last_success_at=now_local(), consecutive_failures=0)
            db.execute(ins.on_conflict_do_update(
                index_elements=["source"],
                set_={"last_success_at": ins.excluded.last_success_at,
                      "last_error": None, "consecutive_failures": 0}))
        if own_session:
            if dry_run:
                db.rollback()
            else:
                db.commit()
        else:
            db.flush()  # 外部会话由调用方（get_db）提交

        totals = {k: ex_stats[k] + pf_stats.get(k, 0)
                  for k in ("rows", "inserted", "failed")}
        # 任务卡「跳过」= 业务 upsert 命中 + 跨源 dup + profile 留档；报告里 dup 单列
        totals["skipped"] = ex_stats["skipped"] + ex_stats["dup_samsung"] + pf_stats["skipped"]
        report["totals"] = totals
        _job_update(job_id, status="done", finished_at=now_local(), report=report,
                    total=totals["rows"], inserted=totals["inserted"],
                    skipped=totals["skipped"], failed=totals["failed"])
        return report
    except Exception as e:
        if own_session and db is not None:
            db.rollback()
        _job_update(job_id, status="failed", finished_at=now_local(),
                    error=str(e)[:2000], report=report)
        raise
    finally:
        if own_session and db is not None:
            db.close()
        shutil.rmtree(workdir, ignore_errors=True)


# ---------- CLI ----------

def main(argv: list[str] | None = None) -> None:
    ap = argparse.ArgumentParser(
        prog="python -m app.importers.keep_file",
        description="Keep 导出文件导入（.7z / .zip / .xlsx，设计文档 §3.7 通道 3）",
    )
    ap.add_argument("path", help="keep 导出文件路径（.7z / .zip / .xlsx）")
    ap.add_argument("--dry-run", action="store_true", help="只解析输出报告，不写库")
    ap.add_argument("--password", default=None, help="压缩包解压密码（可选）")
    args = ap.parse_args(argv)
    report = import_keep(args.path, dry_run=args.dry_run, password=args.password)
    print(json.dumps(report, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
